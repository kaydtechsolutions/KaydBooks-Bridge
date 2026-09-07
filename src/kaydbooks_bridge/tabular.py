"""Bounded spreadsheet intake; explicit mappings prepare drafts, never accounting writes."""

import base64
import copy
import csv
import hashlib
import io
import json
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from defusedxml.ElementTree import fromstring
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from . import documents
from .config import BridgeError, company_policy_context, identifier, strict_keys
from .service import audited, validate_payload
from .validation import canonical, digest

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_ROWS = 1000
MAX_COLUMNS = 64


def schema(db):
    db.execute("""CREATE TABLE IF NOT EXISTS intake_previews (
        id TEXT PRIMARY KEY, owner TEXT NOT NULL, document_id TEXT NOT NULL,
        specification TEXT NOT NULL, plan TEXT NOT NULL)""")
    db.execute("""CREATE TABLE IF NOT EXISTS intake_links (
        preview_id TEXT NOT NULL REFERENCES intake_previews(id), row_key TEXT NOT NULL,
        job_id TEXT NOT NULL REFERENCES jobs(id), PRIMARY KEY(preview_id,row_key))""")
    for table in ("intake_previews", "intake_links"):
        for action in ("UPDATE", "DELETE"):
            db.execute(f"""CREATE TRIGGER IF NOT EXISTS {table}_no_{action.lower()}
                BEFORE {action} ON {table} BEGIN SELECT RAISE(ABORT,'immutable intake evidence'); END""")


def scalar(value):
    if isinstance(value, datetime):
        if value.time().isoformat() != "00:00:00":
            raise BridgeError("spreadsheet date contains a time; use an explicit ISO date")
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    if type(value) in (str, int, float, bool):
        if len(str(value)) > 8192 or (type(value) is float and not Decimal(str(value)).is_finite()):
            raise BridgeError("oversized or nonfinite cell")
        return value
    raise BridgeError("unsupported spreadsheet cell")


def workbook_rows(content, sheet):
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            names = [e.filename for e in entries]
            if (
                len(entries) > 128
                or len(set(names)) != len(names)
                or sum(e.file_size for e in entries) > 16 * 1024 * 1024
                or any(e.flag_bits & 1 for e in entries)
                or any(
                    any(
                        x in name.lower()
                        for x in ("vbaproject", "externallinks", "embeddings", "customui")
                    )
                    for name in names
                )
            ):
                raise BridgeError("unsupported or oversized workbook package")
            for name in names:
                if not name.endswith((".xml", ".rels")):
                    continue
                tree = fromstring(archive.read(name))
                for node in tree.iter():
                    content_type = node.attrib.get("ContentType", "").lower()
                    relation = node.attrib.get("Type", "").rsplit("/", 1)[-1].lower()
                    if any(
                        x in content_type for x in ("macroenabled", "vbaproject", "oleobject")
                    ) or relation in {"externallink", "oleobject", "package"}:
                        raise BridgeError(
                            "macros, external workbook links and embedded objects are unsupported"
                        )
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    previous_row, previous_column = 0, 0
                    for node in tree.iter():
                        tag = node.tag.rsplit("}", 1)[-1]
                        if tag == "mergeCell":
                            raise BridgeError("merged cells are ambiguous; use a flat table")
                        if tag == "row":
                            number = int(node.attrib["r"])
                            if not previous_row < number <= MAX_ROWS + 1:
                                raise BridgeError(
                                    "workbook rows must be ordered, unique and bounded"
                                )
                            previous_row, previous_column = number, 0
                        if tag == "c":
                            col, row = coordinate_from_string(node.attrib["r"])
                            column = column_index_from_string(col)
                            if row != previous_row or not previous_column < column <= MAX_COLUMNS:
                                raise BridgeError(
                                    "workbook cells must be ordered, unique and bounded"
                                )
                            previous_column = column
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
        try:
            if not isinstance(sheet, str) or sheet not in book.sheetnames:
                raise BridgeError("select an exact worksheet name")
            page = book[sheet]
            page.reset_dimensions()
            result = []
            for cells in page.iter_rows():
                errors = []
                values = []
                for cell in cells:
                    if cell.data_type in ("f", "e"):
                        errors.append("formula or error cell requires explicit values")
                    try:
                        values.append(scalar(cell.value))
                    except BridgeError as exc:
                        errors.append(str(exc))
                        values.append("")
                result.append((values, errors))
            return result
        finally:
            book.close()
    except BridgeError:
        raise
    except Exception as exc:
        raise BridgeError("invalid or unsupported XLSX workbook") from exc


def table(content, media_type, specification):
    if media_type == "text/csv":
        if specification.get("sheet") is not None or specification.get("delimiter") not in (
            ",",
            ";",
            "\t",
            "|",
        ):
            raise BridgeError("CSV requires an explicit supported delimiter and no worksheet")
        try:
            text = content.decode("utf-8-sig")
            reader = csv.reader(
                io.StringIO(text, newline=""), delimiter=specification["delimiter"], strict=True
            )
            rows = []
            for values in reader:
                if len(rows) > MAX_ROWS or len(values) > MAX_COLUMNS:
                    raise BridgeError("table row or column limit exceeded")
                errors = []
                for value in values:
                    if len(value) > 8192:
                        errors.append("oversized cell")
                rows.append((values, errors))
        except (UnicodeError, csv.Error) as exc:
            raise BridgeError("CSV must be valid UTF-8 with consistent quoting") from exc
    elif media_type == XLSX:
        if "delimiter" in specification:
            raise BridgeError("XLSX uses a worksheet, not a delimiter")
        rows = workbook_rows(content, specification.get("sheet"))
    else:
        raise BridgeError("intake supports UTF-8 CSV and XLSX only")
    if not rows or len(rows) > MAX_ROWS + 1:
        raise BridgeError("empty or oversized table")
    headers, errors = rows[0]
    # Trailing empty worksheet cells are not additional columns.
    while headers and headers[-1] == "":
        headers.pop()
    if (
        errors
        or not headers
        or len(headers) > MAX_COLUMNS
        or any(not isinstance(h, str) or not 1 <= len(h) <= 128 or h.strip() != h for h in headers)
        or len(set(headers)) != len(headers)
    ):
        raise BridgeError("unique nonempty exact text headers required")
    result = []
    for number, (values, errors) in enumerate(rows[1:], 2):
        if not any(v != "" for v in values) and not errors:
            continue
        if len(values) > len(headers) and any(v != "" for v in values[len(headers) :]):
            errors = [*errors, "row has cells beyond its headers"]
        values = values[: len(headers)] + [""] * max(0, len(headers) - len(values))
        result.append((number, dict(zip(headers, values, strict=True)), errors))
    if not result:
        raise BridgeError("table has no data rows")
    return headers, result


def convert(value, kind):
    if isinstance(value, str) and value.lstrip().startswith(("=", "@")):
        raise BridgeError("formula-like value is not accepted")
    if kind == "text":
        if not isinstance(value, str) or not value or value != value.strip():
            raise BridgeError("exact nonempty text required; format identifiers as text")
        return value
    if kind == "date":
        if not isinstance(value, str) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
            raise BridgeError("date requires YYYY-MM-DD")
        try:
            return date.fromisoformat(value).isoformat()
        except ValueError as exc:
            raise BridgeError("invalid calendar date") from exc
    if kind not in ("money", "decimal") or type(value) not in (str, int, float):
        raise BridgeError("unsupported cell conversion")
    text = str(value)
    if not re.fullmatch(r"-?(?:0|[1-9][0-9]{0,11})(?:\.[0-9]{1,6})?", text):
        raise BridgeError("plain bounded decimal required; currency symbols/grouping are ambiguous")
    try:
        amount = Decimal(text)
        if kind == "money":
            if amount != amount.quantize(Decimal("0.01")):
                raise BridgeError("money requires exact cents; rounding is not implicit")
            return format(amount, ".2f")
        return format(amount, "f")
    except InvalidOperation as exc:
        raise BridgeError("invalid decimal") from exc


def assign(payload, path, value):
    if not isinstance(path, str) or not re.fullmatch(
        r"[a-z_][a-z0-9_]*(?:\.(?:[a-z_][a-z0-9_]*|[0-9]{1,2}))*", path
    ):
        raise BridgeError("explicit bounded template field path required")
    parts = path.split(".")
    current = payload
    try:
        for part in parts[:-1]:
            current = current[int(part)] if isinstance(current, list) else current[part]
        key = int(parts[-1]) if isinstance(current, list) else parts[-1]
        if isinstance(current[key], (dict, list)):
            raise KeyError(path)
        current[key] = value
    except (KeyError, IndexError, ValueError, TypeError) as exc:
        raise BridgeError("mapping must target an existing scalar template field") from exc


def make_plan(policy, content, media_type, specification):
    strict_keys(
        specification,
        {"dataset", "operation", "key_column", "template", "columns"},
        {"sheet", "delimiter"},
    )
    identifier(specification["dataset"])
    if not isinstance(specification["template"], dict) or len(canonical(specification)) > 65536:
        raise BridgeError("bounded mapping template required")
    columns = specification["columns"]
    if not isinstance(columns, dict) or not 1 <= len(columns) <= 256:
        raise BridgeError("explicit column mapping required")
    headers, rows = table(content, media_type, specification)
    if specification["key_column"] not in headers:
        raise BridgeError("select a stable row identity column")
    for path, mapping in columns.items():
        strict_keys(mapping, {"column", "type"})
        if mapping["column"] not in headers or mapping["type"] not in (
            "text",
            "date",
            "money",
            "decimal",
        ):
            raise BridgeError("mapping column or conversion unavailable")
        assign(copy.deepcopy(specification["template"]), path, "")
    keys = [values[specification["key_column"]] for _, values, _ in rows]
    results = []
    for number, values, errors in rows:
        errors = list(errors)
        row_key = values[specification["key_column"]]
        payload = copy.deepcopy(specification["template"])
        if (
            not isinstance(row_key, str)
            or not 1 <= len(row_key) <= 128
            or row_key.strip() != row_key
        ):
            errors.append("stable row identity requires bounded exact text")
        elif keys.count(row_key) != 1:
            errors.append("duplicate row identity in this batch")
        for path, mapping in columns.items():
            try:
                assign(payload, path, convert(values[mapping["column"]], mapping["type"]))
            except BridgeError as exc:
                errors.append(f"{path}: {exc}")
        if not errors:
            try:
                payload = validate_payload(specification["operation"], payload, policy)
            except BridgeError as exc:
                errors.append(str(exc))
        results.append(
            {
                "row": number,
                "row_key": row_key,
                "payload": payload if not errors else None,
                "errors": errors,
                "values": values,
            }
        )
    return {
        "policy": digest(company_policy_context(policy)),
        "rows": results,
        "operation": specification["operation"],
        "source_sha256": hashlib.sha256(content).hexdigest(),
        "prepares_drafts_only": True,
        "master_checks_required_at_prepare": True,
    }


@audited
def preview(bridge, token, company, document_id, specification):
    _, actor, policy, store = bridge._context(token, company, "prepare")
    with store.transaction() as db:
        documents.schema(db)
        schema(db)
        row = db.execute(
            "SELECT * FROM documents WHERE id=? AND owner=?", (document_id, actor)
        ).fetchone()
        if row is None or not store.verify_audit(db):
            raise BridgeError("owned intact source document required")
        plan = make_plan(policy, row["bytes"], row["media_type"], specification)
        preview_id = digest(
            {
                "company": company,
                "owner": actor,
                "document": document_id,
                "specification": specification,
                "plan": plan,
            }
        )
        db.execute(
            "INSERT OR IGNORE INTO intake_previews VALUES (?,?,?,?,?)",
            (preview_id, actor, document_id, canonical(specification), canonical(plan)),
        )
        store.event(
            db,
            bridge.clock(),
            actor,
            None,
            "intake_preview",
            {"preview_id": preview_id, "rows": len(plan["rows"])},
        )
    return {"preview_id": preview_id, "company": company, **plan}


@audited
def prepare_rows(bridge, token, company, preview_id, row_keys, master_evidence=None):
    _, actor, policy, store = bridge._context(token, company, "prepare")
    if (
        not isinstance(row_keys, list)
        or not 1 <= len(row_keys) <= MAX_ROWS
        or any(not isinstance(k, str) for k in row_keys)
        or len(set(row_keys)) != len(row_keys)
    ):
        raise BridgeError("select distinct preview row identities")
    evidence = {} if master_evidence is None else master_evidence
    if not isinstance(evidence, dict) or set(evidence) - set(row_keys):
        raise BridgeError("master evidence must be keyed by selected row identity")
    with store.transaction() as db:
        documents.schema(db)
        schema(db)
        retained = db.execute(
            "SELECT * FROM intake_previews WHERE id=? AND owner=?", (preview_id, actor)
        ).fetchone()
        if retained is None or not store.verify_audit(db):
            raise BridgeError("owned intact preview required")
        spec, plan = json.loads(retained["specification"]), json.loads(retained["plan"])
        original = db.execute(
            "SELECT * FROM documents WHERE id=? AND owner=?", (retained["document_id"], actor)
        ).fetchone()
        if (
            original is None
            or make_plan(policy, original["bytes"], original["media_type"], spec) != plan
        ):
            raise BridgeError("preview or company policy changed; review a fresh preview")
        available = {r["row_key"] for r in plan["rows"] if isinstance(r["row_key"], str)}
        if set(row_keys) - available:
            raise BridgeError("selected row is not in the reviewed preview")
    results = []
    for row in plan["rows"]:
        if row["row_key"] not in row_keys:
            continue
        errors = row["errors"]
        job = None
        if not errors:
            try:
                # Reauthorization and policy checks repeat per row, including after partial restart.
                _, _, current, _ = bridge._context(token, company, "prepare")
                if digest(company_policy_context(current)) != plan["policy"]:
                    raise BridgeError("company policy changed; review a fresh preview")
                row_source = {
                    "dataset": spec["dataset"],
                    "row_key": row["row_key"],
                    "values": row["values"],
                    "operation": spec["operation"],
                    "template": spec["template"],
                    "columns": spec["columns"],
                }
                identity = (
                    "row-" + digest([original["namespace"], spec["dataset"], row["row_key"]])[:59]
                )
                source_ref = "capture-" + digest([identity, row_source])[:55]
                source = documents.capture(
                    bridge,
                    token,
                    company,
                    original["namespace"],
                    source_ref,
                    "application/json",
                    base64.b64encode(canonical(row_source).encode()).decode(),
                )
                job = documents.prepare(
                    bridge,
                    token,
                    company,
                    source["document_id"],
                    identity,
                    row["payload"],
                    dict.fromkeys(documents.fields(row["payload"]), 1),
                    evidence.get(row["row_key"]),
                    operation=spec["operation"],
                )
                with store.transaction() as db:
                    db.execute(
                        "INSERT OR IGNORE INTO intake_links VALUES (?,?,?)",
                        (preview_id, row["row_key"], job["id"]),
                    )
                    store.event(
                        db,
                        bridge.clock(),
                        actor,
                        job["id"],
                        "intake_row_prepared",
                        {
                            "preview_id": preview_id,
                            "row": row["row"],
                            "row_key": row["row_key"],
                            "source_document": retained["document_id"],
                            "source_sha256": plan["source_sha256"],
                        },
                    )
            except BridgeError as exc:
                errors = [str(exc)]
        results.append(
            {
                "row": row["row"],
                "row_key": row["row_key"],
                "errors": errors,
                "job_id": job["id"] if job else None,
                "state": job["state"] if job else None,
            }
        )
    return {"preview_id": preview_id, "company": company, "rows": results, "accounting_writes": 0}
